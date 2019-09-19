import openpyxl
course = 'I:/From Home/Intro to CS/GRADES/2019-09-09T1119_Grades-INTRO_TO_COMPUTER_SCIENCE.xlsx'
wb = openpyxl.load_workbook(filename = course)
ws = wb.active
scores = [[0 for x in range(ws.max_column+1)] for y in range(ws.max_row+1)]
sc=0

def determine_row(r,ws,wb):
    cellOne = ws.cell(row=r,column=1).value
    if  cellOne == "Student":
        return "header\n"
    elif "Points" in cellOne:
        return "total\n"
    else:
        return "body"
def fillAssignments(ws,row):
    print("fill headers")
    unusable = ["Unposted","Final","Current"]
    global headers
    headers = []
    
    for cols in range(1,ws.max_column + 1):
        if not(any(x in ws.cell(row=row, column=cols).value for x in unusable)):
            headers.append(ws.cell(row=row, column=cols).value)
    #print(headers)
    return headers

def addTotals(ws,row):
    print("fill totals")
    unusable = ["Unposted","Final","Current"]
    global totals
    totals = []
    header=1
    
    for cols in range(1,ws.max_column + 1):
        if not(any(x in ws.cell(row=header, column=cols).value for x in unusable)):
            totals.append(ws.cell(row=row, column=cols).value)
    print("totals",totals)
    return totals

def processStudent(ws,row):
    print("fill totals")
    unusable = ["Unposted","Final","Current"]
    global students
    students = []
    header=1
    
    for cols in range(1,ws.max_column + 1):
        if not(any(x in ws.cell(row=header, column=cols).value for x in unusable)):
            students.append(ws.cell(row=row, column=cols).value)
    print("students;",students)
    return students

for row in range(1,ws.max_row + 1):
    print(determine_row(row,ws,wb),end=" ")
    if "header" in determine_row(row,ws,wb):
        headers=fillAssignments(ws,row)
        
    elif "total" in determine_row(row,ws,wb):
        totals=addTotals(ws,row)
    else:
        students=processStudent(ws,row)
        print(students)
        scores[sc]=students
        sc+=1
        print("sc is now: ",sc,"max row:",ws.max_row,"max col",ws.max_column)

for row in range(len(headers)):
    print(headers[row],":",totals[row],":",scores[0][row])



def week(i):    #this is a dictionary
        switcher={
                0:'Sunday',
                1:'Monday',
                2:'Tuesday',
                3:'Wednesday',
                4:'Thursday',
                5:'Friday',
                6:'Saturday'
             }
        return switcher.get(i,"Invalid day of week")
