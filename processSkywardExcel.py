import openpyxl
course = 'I:/From Home/Intro to CS/GRADES/gb_export (25).xlsx'
wb = openpyxl.load_workbook(filename = course)
ws = wb.active
scores = [[0 for x in range(ws.max_column+1)] for y in range(ws.max_row+1)]
sc=0

def determine_row(r,ws,wb):
    #print ("determine",ws.cell(row=r,column=1).value,"row is",r)
    
    cellCheck = str(ws.cell(row=r,column=1).value)+str(ws.cell(row=r,column=2).value)
    #print("checking",cellCheck)
    if  "Last" in cellCheck:
        return "header\n"
    elif "Max" in cellCheck:
        return "total\n"
    elif str(ws.cell(row=r, column=1).value).isnumeric():
        return "body"
    else:
        return "n/a"
def fillAssignments(ws,row):
    print("fill headers")
    unusable = ["Unposted","Final","Current"]
    global headers
    headers = []
    
    for cols in range(1,ws.max_column + 1):
        if not(any(x in ws.cell(row=row, column=cols).value for x in unusable)):
            headers.append(ws.cell(row=row, column=cols).value)
    #print(headers)
    return headers

def addTotals(ws,row):
    print("fill totals")
    unusable = ["Unposted","Final","Current"]
    global totals
    totals = []
    header=1
    
    for cols in range(1,ws.max_column + 1):
        if not(any(x in ws.cell(row=header, column=cols).value for x in unusable)):
            totals.append(ws.cell(row=row, column=cols).value)
    print("totals",totals)
    return totals

def processStudent(ws,row):
    print("fill student")
    unusable = ["Unposted","Final","Current"]
    global students
    students = []
    header=1
    
    for cols in range(1,ws.max_column + 1):
        if not(any(x in str(ws.cell(row=header, column=cols).value) for x in unusable)):
            students.append(str(ws.cell(row=row, column=cols).value))
    print("students;",students)
    return students

for line in range(1,ws.max_row + 1):
    print("line is",line)
    lineType = determine_row(line,ws,wb)
    if "header" in lineType:
        print("checking header")
        headers=fillAssignments(ws,line)
        
    elif "total" in lineType:
        print("checking total")
        totals=addTotals(ws,line)
    elif "body" in lineType:
        print("checking body")
        students=processStudent(ws,line)
        print(students)
        scores[sc]=students
        sc+=1
        print("sc is now: ",sc,"max row:",ws.max_row,"max col",ws.max_column)
    else:
        print("n/a")

for row in range(len(headers)):
    print(headers[row],":",totals[row],":",scores[0][row])



